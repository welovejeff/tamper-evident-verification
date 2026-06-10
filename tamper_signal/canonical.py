"""Canonicalization and semantic hashing.

Two hashes exist per data artifact:

- Evidence hash: SHA-256 of the raw file bytes, computed once at ingest, never
  recomputed downstream. Anchors the original artifact. See `evidence_hash`.
- Semantic hash: SHA-256 of the canonicalized data *content*. Stable across
  format round-trips (xlsx re-save, xlsx -> CSV, xlsx -> JSON) so long as the
  values are unchanged. See `canonicalize` / `semantic_hash`.

The same canonicalization accepts either a parsed worksheet or an in-memory
list-of-dicts, so a transform's output hashes identically whether it lives in
memory or was written to disk and re-parsed.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import re
import unicodedata
from decimal import ROUND_HALF_EVEN, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

# Quantization for all numeric values. Six decimal places, banker's rounding.
_QUANT = Decimal("0.000001")

_WHITESPACE_RUN = re.compile(r"\s+")


# ---------------------------------------------------------------------------
# Evidence hash
# ---------------------------------------------------------------------------
def evidence_hash(raw_bytes: bytes) -> str:
    """SHA-256 of raw file bytes -> lowercase hex. Computed once at ingest."""
    return hashlib.sha256(raw_bytes).hexdigest()


# ---------------------------------------------------------------------------
# Numeric serialization
# ---------------------------------------------------------------------------
def decimal_to_plain_string(value: Decimal) -> str:
    """Serialize a Decimal as plain decimal notation, never scientific.

    Quantizes to six places with ROUND_HALF_EVEN, then strips trailing zeros so
    that 1, 1.0 and "1" all collapse to "1". Integers carry no decimal point.
    `str(Decimal.normalize())` can emit scientific notation (e.g. "1E+2"); this
    helper guards against that by formatting with an explicit exponent of zero.
    """
    quantized = value.quantize(_QUANT, rounding=ROUND_HALF_EVEN)
    # Collapse negative/positive zero to a single "0" so -0.0 and 0.0 (and any
    # value that quantizes to zero) share one canonical form. Without this the
    # sign branch below would emit "-0", giving two representations for zero.
    if quantized == 0:
        return "0"
    # normalize() collapses trailing zeros (1.000000 -> 1) but may yield "1E+2".
    normalized = quantized.normalize()
    sign, digits, exponent = normalized.as_tuple()
    if exponent >= 0:
        # Integer value possibly represented with a positive exponent; rebuild
        # it as a plain integer string to avoid scientific notation.
        mantissa = "".join(str(d) for d in digits)
        plain = mantissa + ("0" * exponent)
        plain = plain.lstrip("0") or "0"
        return ("-" if sign else "") + plain
    # Fractional value: format() with an 'f' presentation type never uses
    # scientific notation. Trailing zeros were already removed by normalize().
    return format(normalized, "f")


def _coerce_decimal(value: Any) -> Decimal | None:
    """Return a Decimal if `value` parses cleanly as a number, else None."""
    if isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        candidate = value
    elif isinstance(value, int):
        candidate = Decimal(value)
    elif isinstance(value, float):
        # Route floats through str() so 0.1 + 0.2 quantizes stably to 0.3.
        candidate = Decimal(str(value))
    elif isinstance(value, str):
        text = value.strip()
        if text == "":
            return None
        try:
            candidate = Decimal(text)
        except InvalidOperation:
            return None
    else:
        return None
    if not candidate.is_finite():
        return None
    return candidate


# ---------------------------------------------------------------------------
# Header normalization
# ---------------------------------------------------------------------------
def normalize_header(header: Any) -> str:
    """Strip, lowercase, collapse internal whitespace runs to a single '_'.

    "Total  Spend (USD)" -> "total_spend_(usd)".
    """
    text = "" if header is None else str(header)
    text = unicodedata.normalize("NFC", text).strip().lower()
    return _WHITESPACE_RUN.sub("_", text)


def normalize_headers(headers: list[Any]) -> list[str]:
    """Normalize a list of headers, rejecting post-normalization duplicates."""
    normalized = [normalize_header(h) for h in headers]
    seen: dict[str, int] = {}
    for name in normalized:
        seen[name] = seen.get(name, 0) + 1
    dupes = sorted(name for name, count in seen.items() if count > 1)
    if dupes:
        raise ValueError(
            "Duplicate headers after normalization: " + ", ".join(dupes)
        )
    return normalized


# ---------------------------------------------------------------------------
# Cell normalization
# ---------------------------------------------------------------------------
def normalize_cell(value: Any) -> str | bool | None:
    """Normalize a single cell to a JCS-ready leaf (string, bool, or None).

    Numbers are quantized and serialized as plain decimal strings; this means a
    leaf that "looks numeric" is indistinguishable from text that parses as a
    number, which is intended (1, 1.0 and "1" must hash identically).
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, dt.datetime):
        # Naive datetimes are assumed UTC. Aware datetimes are converted to UTC.
        if value.tzinfo is not None:
            value = value.astimezone(dt.timezone.utc)
        # Excel stores dates and datetimes as the same serial number, so a
        # date-only value round-trips through xlsx as a midnight datetime. To
        # keep the semantic hash stable across that round-trip, a datetime whose
        # time component is exactly midnight is canonicalized as a date. The
        # tradeoff: a genuine midnight timestamp collides with the bare date.
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%dT%H:%M:%SZ")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float, Decimal)):
        d = _coerce_decimal(value)
        # Non-finite floats fall through to string handling below.
        if d is not None:
            return decimal_to_plain_string(d)
        return _normalize_string(str(value))
    if isinstance(value, str):
        return _normalize_string(value)
    # Fallback for any other openpyxl-returned type: stringify and normalize.
    return _normalize_string(str(value))


def _normalize_string(text: str) -> str:
    """Unicode NFC, strip leading/trailing whitespace. Empty stays empty."""
    return unicodedata.normalize("NFC", text).strip()


# ---------------------------------------------------------------------------
# Records -> canonical bytes
# ---------------------------------------------------------------------------
def _records_to_table(records: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    """Convert a list-of-dicts into (normalized_headers, rows) form.

    Column set is the union of all keys across records, so rows with missing
    keys are filled with None. Headers are normalized and de-duplicated.
    """
    raw_headers: list[Any] = []
    seen_raw: set[str] = set()
    for record in records:
        for key in record.keys():
            marker = normalize_header(key)
            if marker not in seen_raw:
                seen_raw.add(marker)
                raw_headers.append(key)
    headers = normalize_headers(raw_headers)
    # Map each record's keys to their normalized form for lookup.
    rows: list[list[Any]] = []
    for record in records:
        normalized_record = {normalize_header(k): v for k, v in record.items()}
        rows.append([normalized_record.get(h) for h in headers])
    return headers, rows


def canonicalize(records: list[dict[str, Any]]) -> bytes:
    """Canonicalize records (list-of-dicts) to RFC 8785 JCS bytes.

    Steps: normalize headers, normalize each cell by type, sort rows
    lexicographically by their normalized cell strings (so row order is NOT
    part of integrity), build {"headers": [...], "rows": [...]} with headers in
    normalized-alphabetical order, and serialize per JCS.
    """
    headers, rows = _records_to_table(records)
    return canonicalize_table(headers, rows)


def canonicalize_table(headers: list[str], rows: list[list[Any]]) -> bytes:
    """Canonicalize an already-parsed table. `headers` will be re-normalized."""
    headers = normalize_headers(headers)

    # Sort columns into normalized-alphabetical order, carrying cells along.
    order = sorted(range(len(headers)), key=lambda i: headers[i])
    sorted_headers = [headers[i] for i in order]

    normalized_rows: list[list[str | bool | None]] = []
    for row in rows:
        # Pad short rows with None so column indexing is safe.
        padded = list(row) + [None] * (len(headers) - len(row))
        normalized_rows.append([normalize_cell(padded[i]) for i in order])

    # Sort rows lexicographically by the tuple of their normalized cell strings.
    # Booleans/None are mapped to stable sentinel strings purely for ordering;
    # the serialized values keep their true JSON types.
    def sort_key(r: list[str | bool | None]) -> tuple[str, ...]:
        return tuple(_sort_token(cell) for cell in r)

    normalized_rows.sort(key=sort_key)

    document = {"headers": sorted_headers, "rows": normalized_rows}
    return _jcs_serialize(document)


def _sort_token(cell: str | bool | None) -> str:
    """Stable ordering token. Prefixed so types never collide in sort order."""
    if cell is None:
        return "0:"
    if isinstance(cell, bool):
        return "1:" + ("true" if cell else "false")
    return "2:" + cell


# ---------------------------------------------------------------------------
# Minimal RFC 8785 JCS serializer
# ---------------------------------------------------------------------------
def _jcs_serialize(value: Any) -> bytes:
    """Serialize to RFC 8785 JCS: sorted object keys, no whitespace, UTF-8.

    NOTE: by the time we serialize, every leaf value is a string, bool, or
    null. We therefore do NOT implement JCS's number serialization rules
    (ECMAScript Number.prototype.toString); all numbers were already converted
    to plain decimal strings during cell normalization.
    """
    out: list[str] = []
    _jcs_write(value, out)
    return "".join(out).encode("utf-8")


def canonical_json_bytes(value: Any) -> bytes:
    """Public RFC 8785 JCS serialization of an arbitrary JSON-ish object.

    Used to produce the exact bytes that get signed for a receipt body. Accepts
    str, int, bool, None, dict and list leaves (no floats). The byte output is
    reproducible by a matching JavaScript canonicalizer in badge.js.
    """
    return _jcs_serialize(value)


def _jcs_write(value: Any, out: list[str]) -> None:
    if value is None:
        out.append("null")
    elif value is True:
        out.append("true")
    elif value is False:
        out.append("false")
    elif isinstance(value, int):
        # Receipt bodies carry integer leaves (row_count, null counts). An
        # integer's JCS form is its plain decimal text, which matches
        # JavaScript's JSON.stringify exactly, so browser re-verification of
        # signatures reproduces identical bytes. Floats never reach here:
        # data leaves are pre-stringified and totals sums are decimal strings.
        out.append(str(value))
    elif isinstance(value, str):
        out.append(_jcs_string(value))
    elif isinstance(value, dict):
        out.append("{")
        first = True
        for key in sorted(value.keys()):
            if not first:
                out.append(",")
            first = False
            out.append(_jcs_string(key))
            out.append(":")
            _jcs_write(value[key], out)
        out.append("}")
    elif isinstance(value, (list, tuple)):
        out.append("[")
        first = True
        for item in value:
            if not first:
                out.append(",")
            first = False
            _jcs_write(item, out)
        out.append("]")
    else:
        # Should not happen given the leaf-types invariant above.
        raise TypeError(f"JCS cannot serialize value of type {type(value)!r}")


# JSON string escaping per RFC 8785 (which defers to RFC 8259 minimal escaping).
_JCS_ESCAPES = {
    "\\": "\\\\",
    '"': '\\"',
    "\b": "\\b",
    "\f": "\\f",
    "\n": "\\n",
    "\r": "\\r",
    "\t": "\\t",
}


def _jcs_string(text: str) -> str:
    out = ['"']
    for ch in text:
        if ch in _JCS_ESCAPES:
            out.append(_JCS_ESCAPES[ch])
        elif ord(ch) < 0x20:
            out.append(f"\\u{ord(ch):04x}")
        else:
            out.append(ch)
    out.append('"')
    return "".join(out)


# ---------------------------------------------------------------------------
# xlsx loading
# ---------------------------------------------------------------------------
def load_xlsx(path: str, sheet: str | None = None) -> list[dict[str, Any]]:
    """Parse a worksheet into a list-of-dicts keyed by normalized headers.

    First row is treated as headers. Formula cells use the cached computed value
    openpyxl returns with data_only=True; a formula with no cached value becomes
    None (and is counted as a null downstream). Returning list-of-dicts means
    the same `canonicalize` path is used for on-disk and in-memory data.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet] if sheet is not None else workbook.worksheets[0]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []
        headers = normalize_headers(list(header_row))
        records: list[dict[str, Any]] = []
        for raw in rows_iter:
            # Skip fully empty trailing rows that openpyxl sometimes yields.
            if raw is None or all(cell is None for cell in raw):
                continue
            padded = list(raw) + [None] * (len(headers) - len(raw))
            records.append({headers[i]: padded[i] for i in range(len(headers))})
        return records
    finally:
        workbook.close()


def write_xlsx(records: list[dict[str, Any]], path: str, sheet: str = "Sheet1") -> None:
    """Write a list-of-dicts to an xlsx file.

    Column order is the union of keys in first-seen order. Used by the demo to
    persist each stage's output so it can be re-parsed and hashed identically to
    its in-memory form (the round-trip stability property).
    """
    from openpyxl import Workbook

    columns: list[str] = []
    seen: set[str] = set()
    for record in records:
        for key in record.keys():
            if key not in seen:
                seen.add(key)
                columns.append(key)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet
    worksheet.append(columns)
    for record in records:
        worksheet.append([record.get(column) for column in columns])
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


# ---------------------------------------------------------------------------
# Semantic hash
# ---------------------------------------------------------------------------
def semantic_hash(records: list[dict[str, Any]]) -> str:
    """SHA-256 over canonical bytes of records -> lowercase hex."""
    return hashlib.sha256(canonicalize(records)).hexdigest()


def semantic_hash_table(headers: list[str], rows: list[list[Any]]) -> str:
    """SHA-256 over canonical bytes of a parsed table -> lowercase hex."""
    return hashlib.sha256(canonicalize_table(headers, rows)).hexdigest()
